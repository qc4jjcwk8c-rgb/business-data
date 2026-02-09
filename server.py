#!/usr/bin/env python3
"""
Business Data Builder Server
Serves the HTML application and generates Excel files with proper formatting
"""

from http.server import HTTPServer, SimpleHTTPRequestHandler
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import subprocess
import os

class BusinessDataHandler(SimpleHTTPRequestHandler):
    
    def do_GET(self):
        if self.path == '/health' or self.path == '/healthz':
            # Health check endpoint for deployment platforms
            self.send_response(200)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            self.wfile.write(json.dumps({'status': 'healthy'}).encode())
        elif self.path == '/diagnostic' or self.path == '/diagnostic.html':
            # Diagnostic page
            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.send_header('Cache-Control', 'no-cache')
            self.end_headers()
            
            diag_file = os.path.join(os.path.dirname(__file__), 'diagnostic.html')
            if not os.path.exists(diag_file):
                diag_file = 'diagnostic.html'
            
            with open(diag_file, 'rb') as f:
                self.wfile.write(f.read())
        elif self.path == '/test' or self.path == '/test.html':
            # Test page for debugging
            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.send_header('Cache-Control', 'no-cache')
            self.end_headers()
            
            test_file = os.path.join(os.path.dirname(__file__), 'test.html')
            if not os.path.exists(test_file):
                test_file = 'test.html'
            
            with open(test_file, 'rb') as f:
                self.wfile.write(f.read())
        elif self.path == '/' or self.path == '/index.html':
            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.send_header('Cache-Control', 'no-cache')
            self.end_headers()
            
            # Read and serve the HTML file
            html_file = os.path.join(os.path.dirname(__file__), 'index.html')
            if not os.path.exists(html_file):
                html_file = 'index.html'  # Try current directory
            
            with open(html_file, 'rb') as f:
                self.wfile.write(f.read())
        elif self.path == '/sherloc_logo.jpg':
            self.send_response(200)
            self.send_header('Content-type', 'image/jpeg')
            self.send_header('Cache-Control', 'public, max-age=86400')
            self.end_headers()
            
            logo_file = os.path.join(os.path.dirname(__file__), 'sherloc_logo.jpg')
            if not os.path.exists(logo_file):
                logo_file = 'sherloc_logo.jpg'
            
            with open(logo_file, 'rb') as f:
                self.wfile.write(f.read())
        else:
            super().do_GET()
    
    def do_POST(self):
        if self.path == '/generate-excel':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            config = json.loads(post_data.decode('utf-8'))
            
            try:
                # Generate Excel file
                filepath = self.generate_excel(config)
                
                # Note: Formula recalculation happens when user opens file in Excel
                # We don't need the recalc script on deployed environments
                
                # Send file
                self.send_response(200)
                self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition', 'attachment; filename="business_model.xlsx"')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                
                with open(filepath, 'rb') as f:
                    self.wfile.write(f.read())
                
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                error_response = json.dumps({'error': str(e)})
                self.wfile.write(error_response.encode())
                print(f"Error generating Excel: {e}")
        else:
            self.send_response(404)
            self.end_headers()
    
    def do_OPTIONS(self):
        # Handle preflight requests
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def generate_excel(self, config):
        """Generate Excel workbook based on configuration"""
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        # Generate month headers
        month_headers = []
        current_date = datetime.now()
        for i in range(48):
            date = current_date + timedelta(days=30*i)
            month_headers.append(date.strftime('%b %Y'))
        
        # === STAFF TAB ===
        staff_sheet = workbook.create_sheet('Staff')
        
        # Prepare headers
        headers = ['Position', 'Team', 'Type', 'Direct/Overhead']
        
        extra_category = config.get('extraCategory')
        if extra_category and extra_category.get('name'):
            headers.append(extra_category['name'])
        
        headers.append('Annual Salary')
        headers.extend(month_headers)
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = staff_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        staff_sheet.column_dimensions['A'].width = 20
        staff_sheet.column_dimensions['B'].width = 30
        staff_sheet.column_dimensions['C'].width = 12
        staff_sheet.column_dimensions['D'].width = 18
        
        col_offset = 5
        if extra_category and extra_category.get('name'):
            staff_sheet.column_dimensions[chr(64 + col_offset)].width = 18
            col_offset += 1
        
        staff_sheet.column_dimensions[chr(64 + col_offset)].width = 15
        for i in range(48):
            col_letter = chr(64 + col_offset + 1 + i) if col_offset + 1 + i <= 26 else \
                         chr(64 + (col_offset + 1 + i - 1) // 26) + chr(65 + (col_offset + 1 + i - 1) % 26)
            staff_sheet.column_dimensions[col_letter].width = 12
        
        # Create data validations
        selected_teams = config.get('selectedTeams', [])
        team_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(selected_teams)}"',
            allow_blank=False
        )
        staff_sheet.add_data_validation(team_validation)
        
        type_validation = DataValidation(
            type="list",
            formula1='"PAYE,Contract"',
            allow_blank=False
        )
        staff_sheet.add_data_validation(type_validation)
        
        overhead_validation = DataValidation(
            type="list",
            formula1='"OVERHEAD,DIRECT"',
            allow_blank=False
        )
        staff_sheet.add_data_validation(overhead_validation)
        
        extra_validation = None
        if extra_category and extra_category.get('options'):
            options = [opt for opt in extra_category['options'] if opt.strip()]
            if options:
                extra_validation = DataValidation(
                    type="list",
                    formula1=f'"{",".join(options)}"',
                    allow_blank=True
                )
                staff_sheet.add_data_validation(extra_validation)
        
        # Add employee rows
        current_row = 2
        employee_counts = config.get('employeeCounts', {})
        
        for team in selected_teams:
            count = employee_counts.get(team, 0)
            if count == 0:
                continue
            
            team_abbr = ''.join([word[0].upper() for word in team.split()])
            
            for emp_num in range(1, count + 1):
                # Position
                staff_sheet.cell(row=current_row, column=1, value=f'{team_abbr} Employee {emp_num}')
                
                # Team with dropdown
                team_cell = staff_sheet.cell(row=current_row, column=2, value=team)
                team_validation.add(team_cell)
                
                # Type with dropdown
                type_cell = staff_sheet.cell(row=current_row, column=3, value='PAYE')
                type_validation.add(type_cell)
                
                # Direct/Overhead with dropdown
                overhead_cell = staff_sheet.cell(row=current_row, column=4, value='OVERHEAD')
                overhead_validation.add(overhead_cell)
                
                col = 5
                
                # Extra category column
                if extra_category and extra_category.get('name'):
                    extra_cell = staff_sheet.cell(row=current_row, column=col, value='')
                    if extra_validation:
                        extra_validation.add(extra_cell)
                    col += 1
                
                # Annual Salary
                salary_col_letter = chr(64 + col)
                staff_sheet.cell(row=current_row, column=col, value=0)
                staff_sheet.cell(row=current_row, column=col).number_format = '£#,##0'
                col += 1
                
                # Monthly salary columns
                for month_idx in range(48):
                    formula = f'={salary_col_letter}{current_row}/12'
                    cell = staff_sheet.cell(row=current_row, column=col + month_idx, value=formula)
                    cell.number_format = '£#,##0'
                
                current_row += 1
        
        # Add TOTAL row
        total_row = current_row
        total_cell = staff_sheet.cell(row=total_row, column=1, value='TOTAL')
        total_cell.font = Font(bold=True, size=11)
        
        first_month_col = 6
        if extra_category and extra_category.get('name'):
            first_month_col = 7
        
        for month_idx in range(48):
            col_num = first_month_col + month_idx
            if col_num <= 26:
                col_letter = chr(64 + col_num)
            else:
                col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
            
            formula = f'=SUM({col_letter}2:{col_letter}{current_row - 1})'
            cell = staff_sheet.cell(row=total_row, column=first_month_col + month_idx, value=formula)
            cell.font = Font(bold=True)
            cell.number_format = '£#,##0'
        
        # Freeze panes at C2 (freezes columns A & B, and row 1)
        staff_sheet.freeze_panes = 'C2'
        
        # === SALES TAB ===
        sales_sheet = workbook.create_sheet('Sales', 0)  # Make it first tab
        
        sales_model = config.get('salesModel')
        sales_items = config.get('salesItems', [])
        
        if sales_model == 'custom':
            # Custom model: Item, Description 1, Description 2
            custom_headers = ['Item', 'Description 1', 'Description 2'] + month_headers
            for col_idx, header in enumerate(custom_headers, 1):
                cell = sales_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            sales_sheet.column_dimensions['A'].width = 30
            sales_sheet.column_dimensions['B'].width = 25
            sales_sheet.column_dimensions['C'].width = 25
            
            sales_sheet.cell(row=2, column=1, value='[Add your sales items here]')
            sales_sheet.freeze_panes = 'B2'
            
        elif sales_items:
            # Generate model-specific sales tab
            row_num = 1
            
            # REVENUE SECTION HEADER
            revenue_header = sales_sheet.cell(row=row_num, column=1, value='REVENUE')
            revenue_header.font = Font(bold=True, size=12)
            revenue_header.fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
            sales_sheet.merge_cells(f'A{row_num}:D{row_num}')
            row_num += 1
            
            # Headers row
            headers = ['Item', 'Type', 'Unit Price', 'Growth %'] + month_headers
            for col_idx, header in enumerate(headers, 1):
                cell = sales_sheet.cell(row=row_num, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 1
            
            # Set column widths
            sales_sheet.column_dimensions['A'].width = 30
            sales_sheet.column_dimensions['B'].width = 20
            sales_sheet.column_dimensions['C'].width = 18
            sales_sheet.column_dimensions['D'].width = 12
            
            revenue_start_row = row_num
            
            # Pre-calculate where volume section will start
            # After all revenue items + TOTAL row + 2 blanks + VOLUME header + Volume headers
            volume_section_start_row = revenue_start_row + len(sales_items) + 1 + 2 + 1 + 1
            
            # Helper to safely convert to float
            def to_float(val):
                if val is None or val == '':
                    return 0
                return float(val)
            
            # Store item info for VOLUME and COGS sections
            items_info = []
            
            # Add revenue items
            for item in sales_items:
                item_name = item.get('productName') or item.get('serviceName') or item.get('planName') or \
                           item.get('transactionType') or item.get('productLine') or item.get('usageMetric') or \
                           item.get('streamName', 'Item')
                
                # Extract pricing info
                price = to_float(item.get('unitPrice')) or to_float(item.get('pricePerUnit')) or \
                       to_float(item.get('hourlyRate')) or to_float(item.get('monthlyPrice')) or 0
                
                # Extract volume info
                start_val = to_float(item.get('startingUnits')) or to_float(item.get('startingHours')) or \
                           to_float(item.get('startingSubscribers')) or to_float(item.get('startingGMV')) or \
                           to_float(item.get('startingVolume')) or 0
                
                # Extract growth
                growth = to_float(item.get('monthlyGrowth')) or to_float(item.get('growthRate')) or 0
                
                # Store for later sections
                items_info.append({
                    'name': item_name,
                    'model': sales_model,
                    'price': price,
                    'start_volume': start_val,
                    'growth': growth / 100,
                    'item_data': item
                })
                
                sales_sheet.cell(row=row_num, column=1, value=item_name)
                sales_sheet.cell(row=row_num, column=2, value=sales_model.upper())
                sales_sheet.cell(row=row_num, column=3, value=price)
                sales_sheet.cell(row=row_num, column=3).number_format = '£#,##0'
                sales_sheet.cell(row=row_num, column=4, value=growth / 100)
                sales_sheet.cell(row=row_num, column=4).number_format = '0.0%'
                
                # Revenue formulas: will reference volume section
                volume_row_offset = row_num - revenue_start_row
                
                for month_idx in range(48):
                    col_num = 5 + month_idx
                    if col_num <= 26:
                        col_letter = chr(64 + col_num)
                    else:
                        col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
                    
                    # Formula: Volume × Price
                    # Reference the volume cell in the VOLUME section
                    volume_row_ref = volume_section_start_row + volume_row_offset
                    formula = f'={col_letter}{volume_row_ref}*C{row_num}'
                    
                    cell = sales_sheet.cell(row=row_num, column=col_num, value=formula)
                    cell.number_format = '£#,##0'
                
                row_num += 1
            
            # TOTAL REVENUE row
            total_rev_row = row_num
            total_cell = sales_sheet.cell(row=total_rev_row, column=1, value='TOTAL REVENUE')
            total_cell.font = Font(bold=True, size=11)
            
            for month_idx in range(48):
                col_num = 5 + month_idx
                if col_num <= 26:
                    col_letter = chr(64 + col_num)
                else:
                    col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
                
                formula = f'=SUM({col_letter}{revenue_start_row}:{col_letter}{row_num - 1})'
                cell = sales_sheet.cell(row=total_rev_row, column=col_num, value=formula)
                cell.font = Font(bold=True)
                cell.number_format = '£#,##0'
            
            row_num += 2
            
            # ===== VOLUME SECTION =====
            volume_header = sales_sheet.cell(row=row_num, column=1, value='VOLUME (UNITS / SUBSCRIBERS)')
            volume_header.font = Font(bold=True, size=12)
            volume_header.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            sales_sheet.merge_cells(f'A{row_num}:D{row_num}')
            row_num += 1
            
            # Volume headers (no date headers - those are only in REVENUE section)
            vol_headers = ['Item', 'Type', 'Starting Volume', 'Growth %'] + [''] * 48  # Empty headers for month columns
            for col_idx, header in enumerate(vol_headers, 1):
                cell = sales_sheet.cell(row=row_num, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 1
            
            volume_start_row = row_num
            
            # Add volume items
            for info in items_info:
                sales_sheet.cell(row=row_num, column=1, value=info['name'])
                sales_sheet.cell(row=row_num, column=2, value=info['model'].upper())
                sales_sheet.cell(row=row_num, column=3, value=info['start_volume'])
                sales_sheet.cell(row=row_num, column=3).number_format = '#,##0'
                sales_sheet.cell(row=row_num, column=4, value=info['growth'])
                sales_sheet.cell(row=row_num, column=4).number_format = '0.0%'
                
                # Monthly volume calculations - EDITABLE cells with formulas as starting point
                for month_idx in range(48):
                    col_num = 5 + month_idx
                    
                    if sales_model == 'saas':
                        # SaaS: handle churn
                        item_data = info['item_data']
                        churn = float(item_data.get('churnRate', 0)) / 100
                        growth_rate = float(item_data.get('growthRate', 0)) / 100
                        
                        if month_idx == 0:
                            # Month 1: just starting value
                            value = info['start_volume']
                        else:
                            # Previous month adjusted for churn and growth
                            prev_col = chr(64 + col_num - 1) if col_num - 1 <= 26 else \
                                      chr(64 + (col_num - 2) // 26) + chr(65 + (col_num - 2) % 26)
                            formula = f'={prev_col}{row_num}*(1-{churn})*(1+{growth_rate})'
                            cell = sales_sheet.cell(row=row_num, column=col_num, value=formula)
                            cell.number_format = '#,##0'
                            continue
                    else:
                        # Standard growth model
                        if month_idx == 0:
                            value = info['start_volume']
                        else:
                            formula = f'=C{row_num}*POWER(1+D{row_num},{month_idx})'
                            cell = sales_sheet.cell(row=row_num, column=col_num, value=formula)
                            cell.number_format = '#,##0'
                            continue
                    
                    # For month 0, set value directly
                    cell = sales_sheet.cell(row=row_num, column=col_num, value=value)
                    cell.number_format = '#,##0'
                
                row_num += 1
            
            # TOTAL VOLUME row
            total_vol_row = row_num
            total_vol_cell = sales_sheet.cell(row=total_vol_row, column=1, value='TOTAL VOLUME')
            total_vol_cell.font = Font(bold=True, size=11)
            
            for month_idx in range(48):
                col_num = 5 + month_idx
                if col_num <= 26:
                    col_letter = chr(64 + col_num)
                else:
                    col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
                
                formula = f'=SUM({col_letter}{volume_start_row}:{col_letter}{row_num - 1})'
                cell = sales_sheet.cell(row=total_vol_row, column=col_num, value=formula)
                cell.font = Font(bold=True)
                cell.number_format = '#,##0'
            
            row_num += 2
            
            # COGS SECTION HEADER
            cogs_header = sales_sheet.cell(row=row_num, column=1, value='COST OF GOODS SOLD (COGS)')
            cogs_header.font = Font(bold=True, size=12)
            cogs_header.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
            sales_sheet.merge_cells(f'A{row_num}:D{row_num}')
            row_num += 1
            
            # COGS headers (no date headers - those are only in REVENUE section)
            cogs_headers = ['Item', 'Type', 'Unit Cost', 'Growth %'] + [''] * 48  # Empty headers for month columns
            for col_idx, header in enumerate(cogs_headers, 1):
                cell = sales_sheet.cell(row=row_num, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row_num += 1
            
            cogs_start_row = row_num
            
            # Add COGS items - reference volume section
            for idx, info in enumerate(items_info):
                sales_sheet.cell(row=row_num, column=1, value=f"{info['name']} - COGS")
                sales_sheet.cell(row=row_num, column=2, value='COGS')
                
                # Extract cost per unit
                item_data = info['item_data']
                
                # Calculate cost - different models use different fields
                cost = (to_float(item_data.get('costPerUnit')) or 
                       to_float(item_data.get('deliveryCost')) or 
                       to_float(item_data.get('costPerSubscriber')) or 
                       (to_float(item_data.get('materialCost')) + 
                        to_float(item_data.get('laborCost')) + 
                        to_float(item_data.get('overheadCost'))) or 
                       to_float(item_data.get('cost')))
                
                sales_sheet.cell(row=row_num, column=3, value=cost)
                sales_sheet.cell(row=row_num, column=3).number_format = '£#,##0'
                sales_sheet.cell(row=row_num, column=4, value=info['growth'])
                sales_sheet.cell(row=row_num, column=4).number_format = '0.0%'
                
                # COGS formulas: Volume × Cost per unit
                # Reference the corresponding volume row
                volume_row_ref = volume_start_row + idx
                
                for month_idx in range(48):
                    col_num = 5 + month_idx
                    if col_num <= 26:
                        col_letter = chr(64 + col_num)
                    else:
                        col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
                    
                    # Formula: Volume × Cost
                    formula = f'={col_letter}{volume_row_ref}*C{row_num}'
                    cell = sales_sheet.cell(row=row_num, column=col_num, value=formula)
                    cell.number_format = '£#,##0'
                
                row_num += 1
            
            # TOTAL COGS row
            total_cogs_row = row_num
            total_cogs_cell = sales_sheet.cell(row=total_cogs_row, column=1, value='TOTAL COGS')
            total_cogs_cell.font = Font(bold=True, size=11)
            
            for month_idx in range(48):
                col_num = 5 + month_idx
                if col_num <= 26:
                    col_letter = chr(64 + col_num)
                else:
                    col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
                
                formula = f'=SUM({col_letter}{cogs_start_row}:{col_letter}{row_num - 1})'
                cell = sales_sheet.cell(row=total_cogs_row, column=col_num, value=formula)
                cell.font = Font(bold=True)
                cell.number_format = '£#,##0'
            
            row_num += 2
            
            # GROSS PROFIT row
            gross_profit_row = row_num
            gp_cell = sales_sheet.cell(row=gross_profit_row, column=1, value='GROSS PROFIT')
            gp_cell.font = Font(bold=True, size=12, color='00B050')
            gp_cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            
            for month_idx in range(48):
                col_num = 5 + month_idx
                if col_num <= 26:
                    col_letter = chr(64 + col_num)
                else:
                    col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
                
                formula = f'={col_letter}{total_rev_row}-{col_letter}{total_cogs_row}'
                cell = sales_sheet.cell(row=gross_profit_row, column=col_num, value=formula)
                cell.font = Font(bold=True, color='00B050')
                cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                cell.number_format = '£#,##0'
            
            sales_sheet.freeze_panes = 'E3'
        
        else:
            # Empty template if no items
            empty_headers = ['Item'] + month_headers
            for col_idx, header in enumerate(empty_headers, 1):
                cell = sales_sheet.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            sales_sheet.column_dimensions['A'].width = 30
            sales_sheet.cell(row=2, column=1, value='[Add your sales items here]')
            sales_sheet.freeze_panes = 'B2'
        
        # === NON-STAFF TAB ===
        non_staff_sheet = workbook.create_sheet('Non-Staff')
        
        # Headers with Category and Annual Cost columns
        ns_headers = ['Item', 'Category', 'Annual Cost'] + month_headers
        for col_idx, header in enumerate(ns_headers, 1):
            cell = non_staff_sheet.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        non_staff_sheet.column_dimensions['A'].width = 40
        non_staff_sheet.column_dimensions['B'].width = 30
        non_staff_sheet.column_dimensions['C'].width = 15
        for i in range(48):
            col_num = i + 4
            if col_num <= 26:
                col_letter = chr(64 + col_num)
            else:
                col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
            non_staff_sheet.column_dimensions[col_letter].width = 12
        
        # Add selected non-staff items
        non_staff_items = config.get('nonStaffItems', {})
        non_staff_quantities = config.get('nonStaffQuantities', {})
        
        current_ns_row = 2
        
        if non_staff_items:
            for item_key, is_selected in non_staff_items.items():
                if is_selected:
                    category, item = item_key.split('|', 1)
                    quantity = non_staff_quantities.get(item_key, 1)
                    
                    # Remove emoji from category for cleaner display
                    clean_category = category.split(' ', 1)[1] if ' ' in category else category
                    
                    for i in range(1, quantity + 1):
                        # Item name (add quantity suffix if > 1)
                        item_name = f"{item} {i}" if quantity > 1 else item
                        non_staff_sheet.cell(row=current_ns_row, column=1, value=item_name)
                        
                        # Category
                        non_staff_sheet.cell(row=current_ns_row, column=2, value=clean_category)
                        
                        # Annual Cost (default 0)
                        non_staff_sheet.cell(row=current_ns_row, column=3, value=0)
                        non_staff_sheet.cell(row=current_ns_row, column=3).number_format = '£#,##0'
                        
                        # Monthly columns (formula: Annual/12)
                        for month_idx in range(48):
                            formula = f'=C{current_ns_row}/12'
                            cell = non_staff_sheet.cell(row=current_ns_row, column=4 + month_idx, value=formula)
                            cell.number_format = '£#,##0'
                        
                        current_ns_row += 1
            
            # Add TOTAL row
            total_ns_row = current_ns_row
            total_ns_cell = non_staff_sheet.cell(row=total_ns_row, column=1, value='TOTAL')
            total_ns_cell.font = Font(bold=True, size=11)
            
            for month_idx in range(48):
                col_num = 4 + month_idx
                if col_num <= 26:
                    col_letter = chr(64 + col_num)
                else:
                    col_letter = chr(64 + (col_num - 1) // 26) + chr(65 + (col_num - 1) % 26)
                
                formula = f'=SUM({col_letter}2:{col_letter}{current_ns_row - 1})'
                cell = non_staff_sheet.cell(row=total_ns_row, column=col_num, value=formula)
                cell.font = Font(bold=True)
                cell.number_format = '£#,##0'
        else:
            # Empty template
            non_staff_sheet.cell(row=2, column=1, value='[Add your non-staff costs here]')
        
        non_staff_sheet.freeze_panes = 'B2'
        
        # Save to temp file
        filepath = '/tmp/business_model.xlsx'
        workbook.save(filepath)
        return filepath

def run_server(port=8000):
    server_address = ('', port)
    httpd = HTTPServer(server_address, BusinessDataHandler)
    print(f'🚀 Business Data Builder Server')
    print(f'📊 Server running on port {port}')
    print(f'🌐 Access at: http://localhost:{port}')
    print(f'📁 Serving files from: {os.getcwd()}')
    print(f'✅ index.html exists: {os.path.exists("index.html")}')
    print(f'✅ sherloc_logo.jpg exists: {os.path.exists("sherloc_logo.jpg")}')
    print('')
    print('Press Ctrl+C to stop the server')
    print('-' * 50)
    httpd.serve_forever()

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 8000))
    if len(sys.argv) > 1:
        port = int(sys.argv[1])
    run_server(port)
